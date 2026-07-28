from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import psycopg
from openpyxl import load_workbook

from .schema import create_catalog_schema

MODEL_WORKBOOK = "技术架构与开发_角色能力模型.xlsx"
MODEL_CODE = "technical-architecture-development-20260509-v1.0"
MODEL_VERSION = "v1.0"
USAGE_SHEET = "00_使用说明"
L1_SHEET = "01_一级能力总览"
L2_SHEET = "02_职级要求矩阵"
L3_SHEET = "03_三级能力详单"
RESOURCE_SHEET = "04_学习材料"

DOMAINS = {
    "P01": "Data Infra 能力",
    "P02": "AI Infra / Agent 能力",
    "P03": "Coding 能力",
    "C01": "基本办公能力",
    "C02": "沟通协作",
    "C03": "学习创新",
}
SOURCE_L1_CODES = (*DOMAINS, "P04", "P05", "P06")
EXPECTED_L2_COUNTS = {"P01": 10, "P02": 10, "P03": 9, "C01": 7, "C02": 7, "C03": 8}
EXPECTED_L3_COUNTS = {"P01": 82, "P02": 41, "P03": 70, "C01": 42, "C02": 35, "C03": 40}
EXPECTED_START_LEVEL_COUNTS = {
    "P4": 105,
    "P5": 64,
    "P4–P5": 38,
    "P6": 34,
    "P5–P6": 30,
    "P6–P7": 13,
    "P7–P8": 9,
    "P7": 6,
    "P8": 6,
    "P6–P8": 4,
    "P5–P8": 1,
}
EMPTY_L2_CODES = {"P02.07", "P02.08", "P02.09", "P02.10"}
MATERIAL_CODE = re.compile(r"(?:P|C)\d{2}-M\d{3}")


def resolve_workbook_dir(anchor: Path | None = None) -> Path:
    """Find the repository capability-model directory without relying on cwd."""
    current = (anchor or Path(__file__).resolve()).parent
    for _ in range(10):
        candidate = current / "capability-model"
        if (candidate / MODEL_WORKBOOK).is_file():
            return candidate
        if current.parent == current:
            break
        current = current.parent
    raise FileNotFoundError(
        f"Could not find capability-model directory containing {MODEL_WORKBOOK}; "
        f"last searched: {current / 'capability-model'}"
    )


@dataclass(frozen=True)
class ImportReport:
    model_count: int
    l1_count: int
    l2_count: int
    l3_count: int
    resource_count: int
    resource_link_count: int
    unmatched_materials: tuple[str, ...]
    hard_errors: tuple[str, ...] = ()
    added_l2_codes: tuple[str, ...] = ()
    resource_code_changes: tuple[str, ...] = ()
    resource_link_changes: tuple[str, ...] = ()


@dataclass(frozen=True)
class _L1:
    code: str
    name: str
    category: str
    overview: str
    source_row: int


@dataclass(frozen=True)
class _L2:
    code: str
    name: str
    parent_code: str
    descriptions: tuple[str, str, str, str, str]
    source_row: int


@dataclass(frozen=True)
class _L3:
    code: str
    name: str
    parent_code: str
    start_level: str
    materials_text: str
    expected_output: str
    estimated_hours: str
    output_type: str | None
    notes: str | None
    source_row: int


@dataclass(frozen=True)
class _Resource:
    code: str
    name: str
    material_type: str
    source_text: str
    purpose: str
    status: str
    source_row: int


@dataclass(frozen=True)
class _CatalogSource:
    l1: list[_L1]
    l2: list[_L2]
    l3: list[_L3]
    resources: dict[str, _Resource]
    unmatched_materials: tuple[str, ...]


def _required(value: object, label: str, row: int) -> str:
    if value is None or not str(value).strip():
        raise ValueError(f"missing {label} at row {row}")
    return str(value)


def _optional(value: object) -> str | None:
    return None if value is None or not str(value).strip() else str(value)


def _open_workbook(path: Path):
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    missing = {USAGE_SHEET, L1_SHEET, L2_SHEET, L3_SHEET, RESOURCE_SHEET} - set(
        workbook.sheetnames
    )
    if missing:
        workbook.close()
        raise ValueError(f"missing worksheets: {sorted(missing)}")
    return workbook


def _parse_source(workbook_dir: Path) -> _CatalogSource:
    workbook = _open_workbook(workbook_dir / MODEL_WORKBOOK)
    try:
        l1_by_code: dict[str, _L1] = {}
        for row_number, row in enumerate(
            workbook[L1_SHEET].iter_rows(min_row=2, values_only=True), 2
        ):
            code = _required(row[1], "L1 code", row_number)
            item = _L1(
                code,
                _required(row[2], "L1 name", row_number),
                _required(row[0], "L1 category", row_number),
                _required(row[3], "L1 overview", row_number),
                row_number,
            )
            if code in l1_by_code:
                raise ValueError(f"duplicate L1 {code}")
            l1_by_code[code] = item
        if set(l1_by_code) != set(SOURCE_L1_CODES):
            raise ValueError("unexpected L1 source baseline")
        if any(l1_by_code[code].name != name for code, name in DOMAINS.items()):
            raise ValueError("inconsistent enabled L1 name")

        l2_by_code: dict[str, _L2] = {}
        for row_number, row in enumerate(
            workbook[L2_SHEET].iter_rows(min_row=2, values_only=True), 2
        ):
            parent_code = _required(row[1], "L1 code", row_number)
            code = _required(row[3], "L2 code", row_number)
            if parent_code not in DOMAINS or not code.startswith(f"{parent_code}."):
                raise ValueError(f"invalid L2 parent at row {row_number}")
            if _required(row[2], "L1 name", row_number) != DOMAINS[parent_code]:
                raise ValueError(f"inconsistent L1 name at row {row_number}")
            item = _L2(
                code,
                _required(row[4], "L2 name", row_number),
                parent_code,
                tuple(
                    _required(row[index], f"P{index - 1} description", row_number)
                    for index in range(5, 10)
                ),  # type: ignore[arg-type]
                row_number,
            )
            if code in l2_by_code:
                raise ValueError(f"duplicate L2 {code}")
            l2_by_code[code] = item
        if (
            Counter(node.parent_code for node in l2_by_code.values())
            != EXPECTED_L2_COUNTS
        ):
            raise ValueError("unexpected L2 distribution")

        l3_by_code: dict[str, _L3] = {}
        for row_number, row in enumerate(
            workbook[L3_SHEET].iter_rows(min_row=2, values_only=True), 2
        ):
            if row[5] is None or not str(row[5]).strip():
                continue
            parent_code = _required(row[3], "L2 code", row_number)
            code = _required(row[5], "L3 code", row_number)
            if parent_code not in l2_by_code or not code.startswith(f"{parent_code}."):
                raise ValueError(f"invalid L3 parent at row {row_number}")
            item = _L3(
                code,
                _required(row[6], "L3 name", row_number),
                parent_code,
                _required(row[7], "recommended start level", row_number),
                _required(row[8], "materials text", row_number),
                _required(row[9], "expected output", row_number),
                _required(row[10], "estimated hours", row_number),
                _optional(row[11]),
                _optional(row[12]),
                row_number,
            )
            if code in l3_by_code:
                raise ValueError(f"duplicate L3 {code}")
            l3_by_code[code] = item
        if (
            Counter(node.parent_code.split(".")[0] for node in l3_by_code.values())
            != EXPECTED_L3_COUNTS
        ):
            raise ValueError("unexpected L3 distribution")
        if (
            Counter(node.start_level for node in l3_by_code.values())
            != EXPECTED_START_LEVEL_COUNTS
        ):
            raise ValueError("unexpected recommended start level distribution")
        actual_empty_l2 = set(l2_by_code) - {
            node.parent_code for node in l3_by_code.values()
        }
        if actual_empty_l2 != EMPTY_L2_CODES:
            raise ValueError("unexpected L2 without L3")

        resources: dict[str, _Resource] = {}
        for row_number, row in enumerate(
            workbook[RESOURCE_SHEET].iter_rows(min_row=2, values_only=True), 2
        ):
            item = _Resource(
                _required(row[0], "material code", row_number),
                _required(row[1], "material name", row_number),
                _required(row[2], "material type", row_number),
                _required(row[3], "material source", row_number),
                _required(row[4], "material purpose", row_number),
                _required(row[5], "material status", row_number),
                row_number,
            )
            if item.code in resources:
                raise ValueError(f"duplicate material code {item.code}")
            resources[item.code] = item
        if len(resources) != 95:
            raise ValueError("unexpected resource baseline")
        unmatched = _unmatched_materials(list(l3_by_code.values()), resources)
        return _CatalogSource(
            [l1_by_code[code] for code in DOMAINS],
            list(l2_by_code.values()),
            list(l3_by_code.values()),
            resources,
            unmatched,
        )
    finally:
        workbook.close()


def _unmatched_materials(
    l3_nodes: list[_L3], resources: dict[str, _Resource]
) -> tuple[str, ...]:
    unmatched: list[str] = []
    for node in l3_nodes:
        codes = MATERIAL_CODE.findall(node.materials_text)
        residual = MATERIAL_CODE.sub("", node.materials_text).strip("、,，;；/ \t\r\n")
        if residual or any(code not in resources for code in codes):
            unmatched.append(node.materials_text)
    return tuple(unmatched)


def _preflight_existing(
    connection: psycopg.Connection, source: _CatalogSource
) -> tuple[set[str], set[str], set[str], set[str]]:
    """Reject an unsafe upgrade before its transaction changes any catalog rows."""
    model = connection.execute(
        "SELECT id FROM capability_model ORDER BY id LIMIT 1"
    ).fetchone()
    if model is None:
        return set(), set(), set(), set()
    rows = connection.execute(
        """
        SELECT node.code, node.node_type, parent.code
        FROM capability_node AS node
        LEFT JOIN capability_node AS parent ON parent.id = node.parent_node_id
        WHERE node.model_id = %s
        """,
        (model[0],),
    ).fetchall()
    existing_l2 = {code for code, node_type, _ in rows if node_type == "L2"}
    existing_l3_parents = {
        code: parent for code, node_type, parent in rows if node_type == "L3"
    }
    source_l2 = {node.code for node in source.l2}
    source_l3_parents = {node.code: node.parent_code for node in source.l3}
    if existing_l2 and (
        not existing_l2 <= source_l2 or source_l2 - existing_l2 != EMPTY_L2_CODES
    ):
        raise ValueError("unsafe L2 catalog diff; only P02.07–P02.10 may be added")
    if existing_l3_parents and set(existing_l3_parents) != set(source_l3_parents):
        raise ValueError("unsafe L3 catalog diff; L3 codes must remain unchanged")
    changed_parents = {
        code
        for code, parent in existing_l3_parents.items()
        if source_l3_parents.get(code) != parent
    }
    if changed_parents:
        raise ValueError(f"unsafe L3 parent diff: {sorted(changed_parents)}")
    existing_resources = {
        row[0]
        for row in connection.execute("SELECT material_code FROM learning_resource")
    }
    existing_links = {
        f"{l3_code}:{material_code}"
        for l3_code, material_code in connection.execute(
            """
            SELECT node.code, resource.material_code
            FROM capability_node_resource AS link
            JOIN capability_node AS node ON node.id = link.node_id
            JOIN learning_resource AS resource ON resource.id = link.resource_id
            WHERE node.model_id = %s AND node.node_type = 'L3'
            """,
            (model[0],),
        )
    }
    source_links = {
        f"{node.code}:{code}"
        for node in source.l3
        for code in MATERIAL_CODE.findall(node.materials_text)
        if code in source.resources
    }
    return (
        source_l2 - existing_l2,
        existing_resources - set(source.resources),
        set(source.resources) - existing_resources,
        existing_links ^ source_links,
    )


def _upsert_node(
    connection: psycopg.Connection,
    *,
    model_id: int,
    parent_node_id: int | None,
    node_type: str,
    code: str,
    name: str,
    sort_order: int,
    l1_category: str | None = None,
    overview: str | None = None,
    descriptions: tuple[str, str, str, str, str] | None = None,
    l3: _L3 | None = None,
    source_sheet: str,
    source_row: int,
) -> int:
    p4, p5, p6, p7, p8 = descriptions or (None, None, None, None, None)
    row = connection.execute(
        """
        INSERT INTO capability_node (
            model_id, parent_node_id, node_type, code, name, sort_order,
            l1_category, overview, p4_description, p5_description, p6_description,
            p7_description, p8_description, recommended_start_level, materials_text,
            expected_output, estimated_hours, output_type, notes,
            source_workbook, source_sheet, source_row
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s
        )
        ON CONFLICT (model_id, code) DO UPDATE SET
            parent_node_id = EXCLUDED.parent_node_id,
            node_type = EXCLUDED.node_type,
            name = EXCLUDED.name,
            sort_order = EXCLUDED.sort_order,
            l1_category = EXCLUDED.l1_category,
            overview = EXCLUDED.overview,
            p4_description = EXCLUDED.p4_description,
            p5_description = EXCLUDED.p5_description,
            p6_description = EXCLUDED.p6_description,
            p7_description = EXCLUDED.p7_description,
            p8_description = EXCLUDED.p8_description,
            recommended_start_level = EXCLUDED.recommended_start_level,
            materials_text = EXCLUDED.materials_text,
            expected_output = EXCLUDED.expected_output,
            estimated_hours = EXCLUDED.estimated_hours,
            output_type = EXCLUDED.output_type,
            notes = EXCLUDED.notes,
            source_workbook = EXCLUDED.source_workbook,
            source_sheet = EXCLUDED.source_sheet,
            source_row = EXCLUDED.source_row
        RETURNING id
        """,
        (
            model_id,
            parent_node_id,
            node_type,
            code,
            name,
            sort_order,
            l1_category,
            overview,
            p4,
            p5,
            p6,
            p7,
            p8,
            l3.start_level if l3 else None,
            l3.materials_text if l3 else None,
            l3.expected_output if l3 else None,
            l3.estimated_hours if l3 else None,
            l3.output_type if l3 else None,
            l3.notes if l3 else None,
            MODEL_WORKBOOK,
            source_sheet,
            source_row,
        ),
    ).fetchone()
    assert row is not None
    return row[0]


def import_catalog(workbook_dir: Path, connection: psycopg.Connection) -> ImportReport:
    source = _parse_source(workbook_dir)
    create_catalog_schema(connection)
    added_l2, removed_resources, added_resources, link_changes = _preflight_existing(
        connection, source
    )

    with connection.transaction():
        model = connection.execute(
            "SELECT id FROM capability_model ORDER BY id LIMIT 1"
        ).fetchone()
        if model is None:
            model_id = connection.execute(
                """
                INSERT INTO capability_model (
                    code, name, version, source_workbook, source_sheet, source_row
                )
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
                """,
                (
                    MODEL_CODE,
                    "技术架构与开发角色能力模型",
                    MODEL_VERSION,
                    MODEL_WORKBOOK,
                    USAGE_SHEET,
                    20,
                ),
            ).fetchone()[0]
        else:
            model_id = model[0]
            connection.execute(
                """
                UPDATE capability_model
                SET name = %s, version = %s, source_workbook = %s,
                    source_sheet = %s, source_row = %s
                WHERE id = %s
                """,
                (
                    "技术架构与开发角色能力模型",
                    MODEL_VERSION,
                    MODEL_WORKBOOK,
                    USAGE_SHEET,
                    20,
                    model_id,
                ),
            )

        node_ids: dict[str, int] = {}
        for order, node in enumerate(source.l1, 1):
            node_ids[node.code] = _upsert_node(
                connection,
                model_id=model_id,
                parent_node_id=None,
                node_type="L1",
                code=node.code,
                name=node.name,
                sort_order=order,
                l1_category=node.category,
                overview=node.overview,
                source_sheet=L1_SHEET,
                source_row=node.source_row,
            )
        for order, node in enumerate(source.l2, 1):
            node_ids[node.code] = _upsert_node(
                connection,
                model_id=model_id,
                parent_node_id=node_ids[node.parent_code],
                node_type="L2",
                code=node.code,
                name=node.name,
                sort_order=order,
                descriptions=node.descriptions,
                source_sheet=L2_SHEET,
                source_row=node.source_row,
            )
        for order, node in enumerate(source.l3, 1):
            node_ids[node.code] = _upsert_node(
                connection,
                model_id=model_id,
                parent_node_id=node_ids[node.parent_code],
                node_type="L3",
                code=node.code,
                name=node.name,
                sort_order=order,
                l3=node,
                source_sheet=L3_SHEET,
                source_row=node.source_row,
            )

        resource_ids: dict[str, int] = {}
        for resource in source.resources.values():
            row = connection.execute(
                """
                INSERT INTO learning_resource (
                    material_code, name, material_type, source_text, purpose, status,
                    source_workbook, source_sheet, source_row
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (material_code) DO UPDATE SET
                    name = EXCLUDED.name, material_type = EXCLUDED.material_type,
                    source_text = EXCLUDED.source_text, purpose = EXCLUDED.purpose,
                    status = EXCLUDED.status,
                    source_workbook = EXCLUDED.source_workbook,
                    source_sheet = EXCLUDED.source_sheet,
                    source_row = EXCLUDED.source_row
                RETURNING id
                """,
                (
                    resource.code,
                    resource.name,
                    resource.material_type,
                    resource.source_text,
                    resource.purpose,
                    resource.status,
                    MODEL_WORKBOOK,
                    RESOURCE_SHEET,
                    resource.source_row,
                ),
            ).fetchone()
            assert row is not None
            resource_ids[resource.code] = row[0]

        source_l3_ids = [node_ids[node.code] for node in source.l3]
        source_resource_ids = list(resource_ids.values())
        if source_l3_ids and source_resource_ids:
            connection.execute(
                """
                DELETE FROM capability_node_resource
                WHERE node_id = ANY(%s) AND resource_id = ANY(%s)
                """,
                (source_l3_ids, source_resource_ids),
            )
        links = {
            (node_ids[node.code], resource_ids[code])
            for node in source.l3
            for code in MATERIAL_CODE.findall(node.materials_text)
            if code in resource_ids
        }
        for node_id, resource_id in links:
            connection.execute(
                """
                INSERT INTO capability_node_resource (node_id, resource_id)
                VALUES (%s, %s) ON CONFLICT DO NOTHING
                """,
                (node_id, resource_id),
            )

    return ImportReport(
        1,
        len(source.l1),
        len(source.l2),
        len(source.l3),
        len(source.resources),
        len(links),
        source.unmatched_materials,
        added_l2_codes=tuple(sorted(added_l2)),
        resource_code_changes=tuple(sorted((*removed_resources, *added_resources))),
        resource_link_changes=tuple(sorted(link_changes)),
    )


def ensure_catalog_initialized(
    connection: psycopg.Connection, workbook_dir: Path
) -> ImportReport | None:
    create_catalog_schema(connection)
    model = connection.execute(
        "SELECT version, source_workbook FROM capability_model ORDER BY id LIMIT 1"
    ).fetchone()
    if model is None or model[0] != MODEL_VERSION or model[1] != MODEL_WORKBOOK:
        return import_catalog(workbook_dir, connection)
    return None

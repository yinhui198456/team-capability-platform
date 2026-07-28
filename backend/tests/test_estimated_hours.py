from collections import Counter
from hashlib import sha256
from pathlib import Path

from app.planning.hours import parse_estimated_hours, summarize_estimated_hours


def test_parse_estimated_hours_supports_single_decimal_and_ranges() -> None:
    expected = {
        "4": (4.0, 4.0, True, False),
        "4.5": (4.5, 4.5, True, False),
        " 4-6 ": (4.0, 6.0, True, True),
        "4–6": (4.0, 6.0, True, True),
        "4—6": (4.0, 6.0, True, True),
        "4~6": (4.0, 6.0, True, True),
        "4～6": (4.0, 6.0, True, True),
    }

    for raw, (minimum, maximum, valid, is_range) in expected.items():
        parsed = parse_estimated_hours(raw)
        assert (
            parsed.min_hours,
            parsed.max_hours,
            parsed.is_valid,
            parsed.is_range,
        ) == (
            minimum,
            maximum,
            valid,
            is_range,
        )


def test_parse_estimated_hours_supports_hour_suffix() -> None:
    expected = {
        "4h": (4.0, 4.0, True, False),
        "4H": (4.0, 4.0, True, False),
        "4 h": (4.0, 4.0, True, False),
        "8–16h": (8.0, 16.0, True, True),
        "8 - 16 h": (8.0, 16.0, True, True),
        "2-3h": (2.0, 3.0, True, True),
        "4.5h": (4.5, 4.5, True, False),
        "4–6H": (4.0, 6.0, True, True),
    }

    for raw, (minimum, maximum, valid, is_range) in expected.items():
        parsed = parse_estimated_hours(raw)
        assert (
            parsed.min_hours,
            parsed.max_hours,
            parsed.is_valid,
            parsed.is_range,
        ) == (
            minimum,
            maximum,
            valid,
            is_range,
        )


def test_parse_estimated_hours_keeps_raw_text_unchanged() -> None:
    parsed = parse_estimated_hours("4–6h")
    assert parsed.raw == "4–6h"
    assert parsed.is_valid is True


def test_estimated_hours_summary_keeps_ranges_and_marks_unparsed_text() -> None:
    summary = summarize_estimated_hours(["4", "4–6", "约半天", None])

    assert summary == {
        "min_hours": 8.0,
        "max_hours": 10.0,
        "has_values": True,
        "has_unparsed": True,
    }
    assert parse_estimated_hours("4–6").max_hours != 46
    assert parse_estimated_hours("未知文本").is_valid is False


def _find_capability_model_dir() -> Path:
    start = Path(__file__).resolve().parent
    for parent in [start, *start.parents]:
        candidate = parent / "capability-model"
        if candidate.is_dir():
            return candidate
    raise FileNotFoundError("capability-model directory not found")


def test_source_workbook_estimated_hours_are_all_parseable() -> None:
    workbook_path = _find_capability_model_dir() / "技术架构与开发_角色能力模型.xlsx"
    assert workbook_path.is_file(), f"Source workbook not found: {workbook_path}"
    assert sha256(workbook_path.read_bytes()).hexdigest() == (
        "2169dcd4312d37b9b7f171d3c24ef0be10112c4eafc2c15c462da32cb75c26d2"
    )

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True)
    sheet = wb["03_三级能力详单"]
    header = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    column = header.get("预计耗时(h)")
    assert column is not None, "Expected '预计耗时(h)' column in L3 sheet"

    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        l1_code = row[header["一级序号"] - 1]
        l2_code = row[header["二级序号"] - 1]
        l3_code = row[header["三级序号"] - 1]
        raw = row[column - 1]
        if all(
            isinstance(value, str) and value.strip()
            for value in (l1_code, l2_code, l3_code, raw)
        ):
            values.append(raw.strip())

    assert len(values) == 310, f"Expected 310 L3 estimated hours, got {len(values)}"
    formats = Counter(values)
    assert sum(value.lower().endswith("h") for value in values) == 187
    assert {
        "8–16h": formats["8–16h"],
        "4–8h": formats["4–8h"],
        "16–32h": formats["16–32h"],
        "12–24h": formats["12–24h"],
        "2-3h": formats["2-3h"],
    } == {
        "8–16h": 30,
        "4–8h": 22,
        "16–32h": 13,
        "12–24h": 13,
        "2-3h": 9,
    }

    parsed = [parse_estimated_hours(value) for value in values]
    invalid = [value for value in parsed if not value.is_valid]
    assert (
        invalid == []
    ), f"Unparseable estimated hours: {[value.raw for value in invalid]}"

    total_min = sum(value.min_hours or 0 for value in parsed)
    total_max = sum(value.max_hours or 0 for value in parsed)
    assert total_min == 2937.0
    assert total_max == 5221.0
